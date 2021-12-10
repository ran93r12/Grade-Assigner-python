import os
import sys
from PyQt5.QtWidgets import *
import tkinter as tk
from tkinter import *
from tkinter.messagebox import *
from PIL import ImageTk,Image 
from itertools import count
import xlrd
from openpyxl import *
import openpyxl
from openpyxl.styles import Font, Alignment
import pandas as pd
import numpy as np

class ImageLabel(tk.Label):
    """a label that displays images, and plays them if they are gifs"""
    def load(self, im):
        if isinstance(im, str):
            im = Image.open(im)
        self.loc = 0
        self.frames = []

        try:
            for i in count(1):
                self.frames.append(ImageTk.PhotoImage(im.copy()))
                im.seek(i)
        except EOFError:
            pass

        try:
            self.delay = im.info['duration']
        except:
            self.delay = 100

        if len(self.frames) == 1:
            self.config(image=self.frames[0])
        else:
            self.next_frame()

    def unload(self):
        self.config(image="")
        self.frames = None

    def next_frame(self):
        if self.frames:
            self.loc += 1
            self.loc %= len(self.frames)
            self.config(image=self.frames[self.loc])
            self.after(self.delay, self.next_frame)

def Processing():
  root = tk.Tk()
  lbl = ImageLabel(root)
  lbl.pack()
  lbl.load('assets/Process.gif')
  root.after(2000, lambda: root.destroy())
  root.mainloop()


def error(c):
  Tk().withdraw()
  print(showerror("Error Message",c))


def Done(s):
  Tk().withdraw()
  print(showinfo("SuccessFully Completed","Done with "+s))

# ********************************************** Tkinter Window Finished***********************************************************
# ********************************************** Grade Assinger Starts ************************************************************

def Relative(r,Sheet):
  l=[5,25,60,80,95]
  j=[90,80,70,60,50]
  Abs = []
  Final = []
  # row = 0
  # for i in range(2,Sheet.nrows+1):
  #   Marks = Sheet.cell(i-1,r).value
  #   if(type(Marks) == float or type(Marks) == int):
  #     row+=1
  c = 0
  for i in l:
    Percent = round((row * i)/100)                 
    absolute_value = Sheet.cell(Percent,r).value
    Abs.append(absolute_value)
    res = absolute_value - j[c]
    if ( res >= 0):
    	Final.append(j[c])
    	c+=1
    else:
      if (-10 <= res <= -1):
        Final.append(absolute_value)
        c+=1
      else:
        Final.append(j[c]-10)
        c+=1
  return Final



def Grade_Provider(l,Name,r,n,Sheet,file_location):
  wb = load_workbook(file_location)
  ws = wb[Name]
  # style = Font(color='00000000',bold=True,size=11)
  # style1 = Alignment(horizontal="center")
  # wcell1 = ws.cell(1,n)
  # wcell1.value = 'Grade'
  # wcell1.font = style
  for i in range(2,Sheet.nrows+1):
    Marks = Sheet.cell(i-1,r).value
    # print(Marks,end="")
    if (type(Marks) == float):
      if (Marks>=l[0] and Marks<=100):
        wcell = ws.cell(i,n)
        wcell.value = 'EX'
        wcell.alignment = style1
      elif (Marks>=l[1] and Marks<=l[0]+1):
        wcell = ws.cell(i,n)
        wcell.value = 'A'
        wcell.alignment = style1
      elif (Marks>=l[2] and Marks<=l[1]+1):
        wcell = ws.cell(i,n)
        wcell.value = 'B'
        wcell.alignment = style1
      elif (Marks>=l[3] and Marks<=l[2]+1):
        wcell = ws.cell(i,n)
        wcell.value = 'C'
        wcell.alignment = style1
      elif (Marks>=l[4] and Marks<=l[3]+1):
        wcell = ws.cell(i,n)
        wcell.value = 'D'
        wcell.alignment = style1
      elif (Marks>=0 and Marks<=l[4]+1):
        wcell = ws.cell(i,n)
        wcell.value = 'R'
        wcell.alignment = style1
      else:
        print("Something went worng")
        error("Error in GradeAssinging  of sheet "+Name)
        
    else:
      if(type(Marks) == str):
          if (Marks == 'AB'):
            wcell = ws.cell(i,n)
            wcell.value = 'AB'
            wcell.alignment = style1
          elif (Marks == 'MP'):
          	wcell = ws.cell(i,n)
          	wcell.value = 'MP'
          	wcell.alignment = style1
          elif (Marks == 'No Data'):
          	wcell = ws.cell(i,n)
          	wcell.value = 'No Data'
          	wcell.alignment = style1
      else:
        print("Something went worng")
        error("Error in Marks Column in sheet "+Name)
  wb.save(file_location)  
  


def To_Merge(file_location):
  df = pd.concat(pd.read_excel('New_'+file_location, sheet_name=None),sort=False, ignore_index=True)
  writer = pd.ExcelWriter('New_'+file_location, engine='openpyxl', mode='a')
  writer.book = load_workbook('New_'+file_location)
  df.to_excel(writer,sheet_name='Total',index=False)
  writer.save()
  writer.close()





def To_sort(file_location):
  workbook = xlrd.open_workbook(file_location)
  sheet_names = workbook.sheet_names()
  for index in range(len(sheet_names)):
    df2 = pd.read_excel(file_location,sheet_name = sheet_names[index])
    try:
    	col = list(df2)[4]
    	df2[col] = df2[col].astype(object)
    	df2[col] = df2[col].fillna(0)
    	check = df2[col].tolist()
    	if ("AB" in check):
    		df2[col] = df2[col].replace({"AB" : -1})
    	if ("MP" in check):
    		df2[col] = df2[col].replace({"MP" : -2})
    	df2 = df2.loc[pd.to_numeric(df2[col], errors='coerce').sort_values(ascending = False).index]
    	df2[col] = df2[col].apply(np.ceil)
    	df2[col] = df2[col].replace({-1 : "AB"})
    	df2[col] = df2[col].replace({-2 : "MP"})
    	df2[col] = df2[col].replace({0.0 : "No Data"})

    except:
    	error("Error in sheet ::"+sheet_names[index])
    	break
    if index == 0:
      writer = pd.ExcelWriter('Output_'+file_location, engine='xlsxwriter')
      df2.to_excel(writer,sheet_name=sheet_names[index],index=False)
      writer.save()
    else:
      writer = pd.ExcelWriter('Output_'+file_location, engine='openpyxl', mode='a')
      writer.book = load_workbook('Output_'+file_location)
      df2.to_excel(writer,sheet_name=sheet_names[index],index=False)
      writer.save()
      writer.close()
  print("Done with Sorting!!")


def Adding_Cutoffs(l,sheet_names,file_location):
  for index in range(len(sheet_names)):
    df1 = pd.read_excel(file_location,sheet_name = sheet_names[index])
    df1.insert(6,"EX",l[index][0])
    df1.insert(7,"A",l[index][1])
    df1.insert(8,"B",l[index][2])
    df1.insert(9,"C",l[index][3])
    df1.insert(10,"D",l[index][4])
    if index == 0:
      writer = pd.ExcelWriter('New_'+file_location, engine='xlsxwriter')
      df1.to_excel(writer,sheet_name=sheet_names[index],index=False)
      writer.save()
    else:
      writer = pd.ExcelWriter('New_'+file_location, engine='openpyxl', mode='a')
      writer.book = load_workbook('New_'+file_location)
      df1.to_excel(writer,sheet_name=sheet_names[index],index=False)
      writer.save()
      writer.close()
  print("Finally done with Grade assigning!")


def ToExecute(file_location):
  try:
  	To_sort(file_location)
  except:
    error("Error while executing Sorting Function.")
  C = []
  workbook = xlrd.open_workbook('Output_'+file_location)
  sheet_names = workbook.sheet_names()
  for index in range(len(sheet_names)):
    Sheet = workbook.sheet_by_index(index)
    try:
      Final = Relative(4,Sheet)
    except:
      error("Error in Finding Absolute Grades")
    C.append(Final)
    print(Final)
    Name = sheet_names[index]
    try:
      Grade_Provider(Final,Name,4,6,Sheet,'Output_'+file_location)
    except:
      error("Error while Assigning Grades")
    print("Done with the sheet :: ",sheet_names[index])
  try:
    Adding_Cutoffs(C,sheet_names,'Output_'+file_location)
    To_Merge('Output_'+file_location)
    os.remove('Output_'+file_location)
  except:
    error("Error in Adding Cutoffs")






# ************************************************ Grade Assinger Ends ************************************************************


def openFileDialog():
  option=QFileDialog.Options()
  file=QFileDialog.getOpenFileName(widget,"Open Single File","Default.xlsx","xlsx Files (*.xlsx)",options=option)
  FileNames = [os.path.basename(file[0])]
  print(FileNames)
  for i in FileNames:
    Processing()
    ToExecute(i)
    Done(i)
  sys.exit(app.exec_())

def openMultiFile():
  option=QFileDialog.Options()
  option|=QFileDialog.DontUseNativeDialog
  file=QFileDialog.getOpenFileNames(widget,"Select Multi File","default.xlsx","xlsx Files (*.xlsx)",options=option)
  FileNames = [os.path.basename(i) for i in file[0]]
  print(FileNames)
  Processing()
  for i in FileNames:
    ToExecute(i)
    Done(i)
  sys.exit(app.exec_())


app=QApplication(sys.argv)

widget=QWidget()
widget.resize(300,300)
widget.setWindowTitle('Grade_Allocator')

button_single_file_dialog=QPushButton("Select Single File")
button_single_file_dialog.clicked.connect(openFileDialog)

button_multi_file=QPushButton("Select Multiple file Files")
button_multi_file.clicked.connect(openMultiFile)

vb = QVBoxLayout(widget)
vb.addWidget(button_single_file_dialog)
vb.addWidget(button_multi_file)

widget.show()

sys.exit(app.exec_())


